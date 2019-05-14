# -*- coding: utf-8 -*-
"""
Created on Tue May 14 20:21:04 2019

@author: Prince Farrar
"""

from datetime import datetime
import pandas as pd
from pandas import ExcelWriter
import os
#import shutil
import sys
import openpyxl
#import xlrd

class ReportDownload:
    
    # Declaring country variable
    countries = []
    countries_CORE_TEAM = []
    countries_IN = []
    todays_date = datetime.now
    
    # Initializing variables within constructor
    def __init__(self):
        self.countries = ["AE", "AO", "BH", "CI", "CM", "DF", "GM", "HK", "IQ", "JO", "MU", "NG", "OM", "QA", "SG", "SL", "SP", "UK", "US", "ZW"]
        self.countries_CORE_TEAM = ["BW", "KE", "TZ", "UG", "ZM", "ZA","GH"]
        self.countries_IN = ["IN"]
        self.todays_date = datetime.now().strftime("%Y%m%d")

    def downloadReport(self, report_name):
        sourcePath = "D:/Suzannah/HK.xls"
        if os.path.isfile(sourcePath) == False:
            print("No files")
            
        #for country in self.countries:
          #  sourcePath = "";
           # if report_name == "2": 
               
                #sourcePath = "/root/Trade DTPRGBO reports/GSSC/" + country + "/Daily/" + self.todays_date + "/" + country + "_D(17) Guarantee Outstanding Claims.xls"
                #if os.path.isfile(sourcePath) == False:
                 #   sourcePath = "/root/Trade DTPRGBO reports/GSSC/" + country + "/Daily/" + self.todays_date + "/" + country + "_(17) Guarantee Outstanding Claims.xls"
                #if os.path.isfile(sourcePath) == False:
                 #   sourcePath = "/root/Trade DTPRGBO reports/GSSC/" + country + "/Daily/" + self.todays_date + "/" + self.todays_date + "/" + country + "_D(17) Guarantee Outstanding Claims.xls"
            
        try:
                                    
            #book = openpyxl.load_workbook(sourcePath)
            #sheet = book.active
            #print(sheet['A1'])
            
            excel_report = pd.read_excel(sourcePath, sheetname=0)
            
            
            excel_report = excel_report.drop(0)
            #excel_report = excel_report.drop(0)
            #print(excel_report)
            headers = excel_report.iloc[0]
            
            #headers = excel_report.iloc[0]
            excel_report = pd.DataFrame(excel_report.values[1:], columns = headers)
            #print(len(excel_report.columns))
            #excel_df = pd.DataFrame(excel_report.values[1:], columns  = headers)
            #print(excel_df)
            
            
            
            #print(excel_report)
            
            #for x in range(0, len(excel_report)):
            #    if excel_report.at[x,'STEP_NO'] != 11 and excel_report.at[x,'STEP_NO'] != 1 and excel_report.at[x,'STEP_NO'] != 2:
            #        excel_report = excel_report.drop(excel_report.at[x,'STEP_NO'])
            
            #print(excel_report) 
            
            destPath = "D:/Suzannah/HK.xlsx"
            with ExcelWriter(destPath) as writer:
                excel_report.to_excel(writer)
                writer.save()
                
            book = openpyxl.load_workbook(destPath)
            sheet = book.active
            sheet.delete_cols(1)
            
            new_headers = ["BRANCH_CODE","DEAL_NO", "CUST_ID","CUST_NAME","NAME_2","STEP_NO","STEP_DATE","EXPIRY_DATE","EXPIRY_DATE","MARGIN_NUMBER","CCY","ÄMOUNT","BALANCE","DEAL_CCY","DEAL_AMOUNT","LIABILITY_BALANCE","SEC_ID","CLAIM_EXPIRY_DATE","TENOR DAYS","EXPOSURE_END_DATE","USD_MARGIN_BAL","RELEASER","MARGIN_TYPE_DESC","MARGIN_MATURITY_DATE","LIEN_PARTY_CODE","LOCATION", "SECURITY_PERFECT","PRODUCT","FINAL_EXPIRY_DATE","COUNTRY","COMMENTS"]
            
            for c in range(1,len(new_headers)+1):
                sheet.cell(row=1,column=c).value = new_headers[c-1]
                
            book.save(destPath)            
                
        except Exception:
            print("error - ", sys.exc_info())
            


obj = ReportDownload()
obj.downloadReport("2")